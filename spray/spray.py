#!/usr/bin/python
#

import argparse
import logging
import os.path
import sys
import yaml

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from pathlib import Path

class Spray():
    """This class handles all 'spreadsheet' aspects.
    """
    def __init__(self, yaml_file):
        self.yaml_file = yaml_file
        self.empty = True
        if os.path.exists(self.yaml_file):
            with open(self.yaml_file, 'r') as yfh:
                self.yct = yaml.safe_load(yfh)
            for key,value in self.yct[0].items():
                if key != value:
                    self.empty = False
                    break
        else:
            self.yct = None

    def _check_path(self, path_to_check):
        """checks if a path exist and if it does not exist, create it.

        Args:
            path_to_check (string): the path to check and create
        """
        the_path = Path(path_to_check)
        if not the_path.exists():
            the_path.mkdir()


    def _generate_header(self, ws, data_set):
        column = ord("A")
        for key,value in data_set.items():
            ws[f"{chr(column)}1"] = key
            column += 1

    def _get_header(self, ws):
        column = 1
        data_set = {}
        self.header = []
        while True:
            cell = ws.cell(row=1,column=column)
            if cell.value == None:
                break
            self.header.append(cell.value)
            column += 1


    def create(self, header):
        """Create a 'table' by creating the very first yaml entry as table header.

        Args:
            header (list): list with header content
        """
        yaml_file = Path(self.yaml_file)

        if yaml_file.exists():
            sys.exit(f"YAYML file {self.yaml_file} already exists")
        # check if path to yaml file already exists
        self._check_path(yaml_file.parent)

        tbl_header = [{}]
        for item in header.split(","):
            tbl_header[0][item]=item
        self.yct = yaml.dump(tbl_header)
        with open(self.yaml_file, "w") as yfh:
            yaml.dump(tbl_header, yfh, sort_keys=False)


    def yaml2xls(self, xls_file):
        """Create a spreadsheet file from a yaml input file.

        Args:
            yaml_file (string): the yaml input file
            xls_file (string): the spreadsheet output (Excel-) file
        """
        # read yaml content
        with open(self.yaml_file, 'r') as fh:
            content = yaml.safe_load(fh)
        # Create a workbook
        wb = Workbook()
        # Get the active worksheet or create a new sheet
        ws = wb.active
        row = 0
        for set in content:
            if row == 0:
                self._generate_header(ws, set)
                row = 2
                continue
            column = 1
            for key,value in set.items():
                cell = ws.cell(row=row, column=column)
                if value is not None:
                    if isinstance(value, str):
                        value = value.rstrip("\n")
                        if value.startswith("-"):
                            value = "'"+value
                    cell.value = value
                    if isinstance(value, str) and "\n" in value:
                        cell.alignment = Alignment(wrap_text=True)
                column += 1
            row += 1

        # Save the workbook to a file
        self._check_path(Path(xls_file).parent)
        wb.save(xls_file)

    def xls2yaml(self, xls_file):
        wb = load_workbook(filename = xls_file)
        # Get the active worksheet
        ws = wb.active
        self.yct = []
        self._get_header(ws)
        row = 2
        while True:
            column = 1
            cell = ws.cell(row=row, column=column)
            if cell.value is None:
                break
            yaml_dataset = {}
            for column in range(0, len(self.header)):
                cell = ws.cell(row=row, column=column+1)
                yaml_dataset[self.header[column]] = cell.value
            row += 1
            self.yct.append(yaml_dataset)
        with open(self.yaml_file, "w") as yfh:
            yaml.dump(self.yct, yfh, sort_keys=False)

    def add(self, data_set):
        ds = data_set.split(',')
        if len(ds) != len(self.yct[0]):
            sys.exit(f"Wrong length of input data: {data_set}")
        data_row = {}
        for index,key in enumerate(self.yct[0].keys()):
            data_row[key] = ds[index]
        if self.empty:
            self.yct[0] = data_row
            self.empty = False
        else:
            self.yct.append(data_row)
        with open(self.yaml_file, "w") as yfh:
            yaml.dump(self.yct, yfh, sort_keys=False)


def main():
    """main function of spray project.

    The main function handles the command line arguments.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-d", "--debug", action="store_true", dest="debug", default=False,
        help="show debug output at console"
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", dest="verbose", default=False,
        help="show more console output"
    )
    parser.add_argument(
        'input',
        nargs="*",
        help="give the task and input for the task. Tasks are:" \
        "- yaml2xls\t convert a yaml-file to a spreadsheet file (Excel)" \
        "- xls2yaml\t convert a spreadsheet file (Excel) to a yaml file" \
        "- new\t create a new yaml file" \
        "See 'spray help <task> to get help for each command."
    )
    parser.add_argument(
        "-i","--interactive",
        action="store_true",
        default=False,
        help="do it in interactive mode.",
    )
    parser.add_argument(
        "-x","--xls_file",
        help="Spreadsheet (Excel) file.",
    )
    parser.add_argument(
        "-y","--yaml_file",
        help="YAML based input file.",
    )
    # Parser arguments
    args = parser.parse_args()

    if args.verbose:
        logging.basicConfig(level=logging.INFO)
    elif args.debug:
        logging.basicConfig(level=logging.DEBUG)

    spray = Spray(args.yaml_file)

    if args.input[0] == "new":
        if len(args.input) > 1:
            spray.create(args.input[1])
        else:
            parser.print_help(file=sys.stderr)
            sys.exit(2)
    elif args.input[0] == "xls2yaml":
        spray.xls2yaml(args.xls_file)
    elif args.input[0] == "yaml2xls":
        spray.yaml2xls(args.xls_file)
    elif args.input[0] == "add":
        spray.add(args.input[1])


if __name__ == '__main__':
    main()